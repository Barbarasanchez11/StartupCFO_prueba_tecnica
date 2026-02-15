import openpyxl
from openpyxl.styles import PatternFill
import os
from src.config import OUTPUT_FILE, INPUT_PL_FILE, INPUT_PL_COLS

def save_to_excel(classified_df, template_path, input_df=None):
    """
    Open the original Excel, find the END row, and insert new data with styling.
    If input_df is provided, also rewrite existing rows to fix corrupted values.
    """
    if classified_df is None or len(classified_df) == 0:
        print("[INFO] No data to write.")
        return

    # 1. Abro el Excel original que me sirve de plantilla

    print(f"[INFO] Opening template: {template_path}")
    wb = openpyxl.load_workbook(template_path, data_only=True, keep_vba=False)
    sheet = wb.active

    # 2. Busco TODAS las filas 'END' para encontrar la PRIMERA (donde empiezan los datos nuevos)
    end_rows = []
    # Miro en la columna 1 (Nº Asiento) que es donde suele estar el END
    for row in range(1, sheet.max_row + 1):
        cell_val = str(sheet.cell(row=row, column=1).value).strip().upper()
        if cell_val == 'END':
            end_rows.append(row)
    
    # Si no encuentro ninguna, escribo al final
    if not end_rows:
        end_row = sheet.max_row + 1
        first_end_row = end_row
        print("[WARNING] Could not find 'END' row. Writing at the end of the sheet.")
    else:
        # Uso la PRIMERA fila END como punto de inserción (donde empiezan los datos nuevos)
        first_end_row = min(end_rows)
        end_row = first_end_row
        
        # Si hay múltiples filas END, elimino las intermedias y la última, dejando solo la primera
        if len(end_rows) > 1:
            print(f"[INFO] Found {len(end_rows)} 'END' rows. Removing all except the first one at row {first_end_row}.")
            # Elimino todas las filas END excepto la primera (de mayor a menor para no afectar los índices)
            for row_to_delete in sorted(end_rows[1:], reverse=True):
                sheet.delete_rows(row_to_delete)
            print(f"[INFO] Cleaned up. Using 'END' at row {end_row} as insertion point.")
        else:
            print(f"[INFO] Found 'END' at row {end_row}. Inserting {len(classified_df)} rows...")
    
   
    if input_df is not None and first_end_row > 2:  # first_end_row > 2 porque fila 1 es header
        print(f"[INFO] Fixing existing rows (1 to {first_end_row-1}) from normalized DataFrame...")
        # Reescribir filas existentes desde el DataFrame normalizado (que tiene valores correctos)
        # IMPORTANTE: Solo reescribir hasta la PRIMERA fila END, y excluir filas END del DataFrame
        for df_idx, (_, row_data) in enumerate(input_df.iterrows(), start=2):  # start=2 porque fila 1 es header
            # No reescribir si es una fila END o si está después de la primera END
            if df_idx >= first_end_row:
                break
            # No reescribir filas END del DataFrame
            if str(row_data.get('Nº Asiento', '')).strip().upper() == 'END':
                continue
            for col_idx, col_name in enumerate(INPUT_PL_COLS, start=1):
                if col_name in row_data:
                    cell_value = row_data[col_name]
                    cell = sheet.cell(row=df_idx, column=col_idx)
                    
                    # Aplicar formatos correctos a filas existentes también
                    if col_name == 'Fecha' and hasattr(cell_value, 'to_pydatetime'):
                        cell.value = cell_value.to_pydatetime()
                        cell.number_format = 'DD/MM/YYYY'
                    elif col_name == 'Mes':
                      
                        cell.number_format = '@'
                        cell.value = str(cell_value) if cell_value else ""
                    elif col_name in ['Debe', 'Haber', 'Saldo', 'Neto']:
                        cell.value = cell_value
                        cell.number_format = '#,##0.00'
                    else:
                        cell.value = cell_value

   
    # 3. Insertar los nuevos datos antes de la primera fila END
    # Esto desplazará la primera END hacia abajo
    sheet.insert_rows(end_row, amount=len(classified_df))
    
    # 4. Después de insertar, eliminar la primera END que ahora está en medio
    # (está en end_row + len(classified_df) porque se desplazó)
    end_row_after_insert = end_row + len(classified_df)
    # Verificar que efectivamente hay una END ahí
    if end_row_after_insert <= sheet.max_row:
        cell_val_after = str(sheet.cell(row=end_row_after_insert, column=1).value).strip().upper()
        if cell_val_after == 'END':
            print(f"[INFO] Removing intermediate 'END' row at {end_row_after_insert} (now in the middle after insertion).")
            sheet.delete_rows(end_row_after_insert)

    # 4. Empiezo a rellenar SOLO las nuevas filas insertadas (no toco las existentes)
    # Creo un estilo amarillo para resaltar los que tienen poca confianza
    warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for i, (index, row_data) in enumerate(classified_df.iterrows()):
        current_row = end_row + i
        
        # Voy columna por columna según mi configuración
        for col_idx, col_name in enumerate(INPUT_PL_COLS, start=1):
            if col_name in row_data:
                cell_value = row_data[col_name]
                
                # Si es una fecha la convierto para que Excel no se vuelva loco
                if col_name == 'Fecha' and hasattr(cell_value, 'to_pydatetime'):
                    cell_value = cell_value.to_pydatetime()
                
            
                cell = sheet.cell(row=current_row, column=col_idx)
                
                # Formateo específico según el tipo de columna (SOLO para nuevas filas)
                if col_name == 'Fecha':
                    cell.value = cell_value
                    cell.number_format = 'DD/MM/YYYY'
                
                elif col_name == 'Mes':
                    
                    # Excel interpreta "abr/25" como fecha, necesitamos forzar texto
                    # IMPORTANTE: Solo aplicamos esto a las NUEVAS celdas, no tocamos las existentes
                    cell.number_format = '@'  # @ es el código de formato texto en Excel
                    cell.value = str(cell_value) if cell_value else ""
                
                elif col_name in ['Debe', 'Haber', 'Saldo', 'Neto']:
                    cell.value = cell_value
                    cell.number_format = '#,##0.00'
                else:
                    # Para otras columnas, simplemente escribir el valor
                    cell.value = cell_value
                
                # Si la confianza es bajita, lo pinto de amarillo para avisar
                if row_data.get('Confidence', 100) < 80:
                    cell.fill = warning_fill

    # 5. Guardo el archivo final
    # Primero me aseguro de que la carpeta de salida existe
    output_dir = os.path.dirname(OUTPUT_FILE)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    print(f"[INFO] Saving results to: {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)
    print("[SUCCESS] Process completed! Check the output folder.")
